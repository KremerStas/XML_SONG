from models import Song
import random
import openpyxl


avtor = input().split(', ')
pesni = []
wb = openpyxl.Workbook()
sheet = wb.active
cell = sheet.cell(row=1, column=1)
cell.value = 'ARTIST'
cell_0 = sheet.cell(row=1, column=2)
cell_0.value = 'TITLE'
n = 0
m = 2
for zapros in avtor:
    command = zapros.split('-')
    for art in Song.select().where(Song.artist == command[0]):
        pesni.append(art.song)
    if pesni:
        first_cell = sheet.cell(row=m, column=1)
        first_cell.value = command[0]
        for i in range(int(command[1])):
            second_cell = sheet.cell(row=n + 2, column=2)
            second_cell.value = random.choice(pesni)
            wb.save('SONG.xml')
            n += 1
            m += 1
    else:
        print('Исполнитель не найден')
    pesni.clear()

